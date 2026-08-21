from __future__ import annotations

import hashlib
import json
import shutil
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw

from app.evidence.catalog import clear_catalog_cache
from app.rag.experimental_assets_batch2 import (
    build_batch2_assets,
    required_sources as batch2_required_sources,
    write_batch2_documents,
)


PROJECT_ROOT = Path(__file__).resolve().parents[3]
CORPUS_ROOT = PROJECT_ROOT / "knowledge_corpus"
TEAM_ROOT = CORPUS_ROOT / "raw_docs" / "team_tire"
ASSET_ROOT = CORPUS_ROOT / "experimental_assets"
MEDIA_ROOT = ASSET_ROOT / "media"
CATALOG_PATH = ASSET_ROOT / "catalog.json"
REGISTRY_PATH = CORPUS_ROOT / "source_registry.json"
SERVICE_PARSED_PATH = CORPUS_ROOT / "parsed_docs" / "team_tire_lyq_service_tire_cases.md"

SERVICE_SOURCE_ID = "team_tire_lyq_service_tire_cases"
THERMAL_SOURCE_IDS = [
    "team_tire_tire_exp_rxq_aging_001",
    "team_tire_rxq_aging_multimodal_evidence",
]
REVIEW_DATE = date(2026, 8, 4).isoformat()

LYQ_SOURCE_ROOT = Path(r"E:\梁永琦\轮胎")
RXQ_IMAGE_WORKBOOK = TEAM_ROOT / "data" / "tire_rxq_damage_image_annotations.xlsx"
RXQ_EXPERIMENT_WORKBOOK = TEAM_ROOT / "data" / "tire_rxq_aging_experiment_records.xlsx"

SERVICE_CASES = [
    ("1-1", "1.1.tif", "15x6.0-6", "严重", "局部磨损严重"),
    ("2-4", "2.4.tif", "15x6.0-6", "一般", "偏磨"),
    ("2-5", "2.5.tif", "15x6.0-6", "中等", "偏磨"),
    ("2-6", "2.6.tif", "15x6.0-6", "严重", "偏磨"),
    ("3-1", "3.1.tif", "15x6.0-6", "一般", "偏磨"),
    ("3-2", "3.2.tif", "15x6.0-6", "中等", "偏磨"),
    ("3-3", "3.3.tif", "15x6.0-6", "严重", "偏磨"),
    ("4-1", "4-1.tif", "22x7.75-10", "未分级", "偏磨"),
    ("4-2", "4.2.tif", "22x7.75-10", "未分级", "偏磨"),
    ("4-3", "4.3.tif", "22x7.75-10", "未分级", "偏磨"),
]


def main() -> None:
    _assert_sources()
    _reset_media()
    service_assets = _build_service_tire_assets()
    thermal_assets = _build_thermal_assets()
    uv_assets, load_assets = build_batch2_assets(_make_thumbnail, _sha256)
    assets = service_assets + thermal_assets + uv_assets + load_assets
    payload = {
        "schema_version": "1.0",
        "status": "ready",
        "generated_at": REVIEW_DATE,
        "source_policy": "Only reviewed derivatives are copied; full-resolution raw data remains on authorized source media.",
        "topics": [
            {
                "topic": "service_tire_wear",
                "title": "实胎偏磨与严重度判读",
                "asset_count": len(service_assets),
                "review_status": "reviewed",
            },
            {
                "topic": "thermal_aging_multimodal",
                "title": "热氧老化多表征证据链",
                "asset_count": len(thermal_assets),
                "review_status": "reviewed",
            },
            {
                "topic": "uv_aging_wear_morphology",
                "title": "紫外老化后摩擦形貌证据",
                "asset_count": len(uv_assets),
                "review_status": "reviewed",
            },
            {
                "topic": "load_distance_morphology_matrix",
                "title": "载荷-距离-形貌训练矩阵",
                "asset_count": len(load_assets),
                "review_status": "reviewed",
            },
        ],
        "assets": assets,
    }
    ASSET_ROOT.mkdir(parents=True, exist_ok=True)
    CATALOG_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_service_markdown(service_assets)
    _register_service_source()
    write_batch2_documents(uv_assets, load_assets, _sha256)
    clear_catalog_cache()
    print(
        json.dumps(
            {
                "status": "prepared",
                "catalog": str(CATALOG_PATH),
                "service_tire_assets": len(service_assets),
                "thermal_aging_assets": len(thermal_assets),
                "uv_aging_wear_assets": len(uv_assets),
                "load_distance_matrix_assets": len(load_assets),
                "total_assets": len(assets),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def _assert_sources() -> None:
    required = [
        RXQ_IMAGE_WORKBOOK,
        RXQ_EXPERIMENT_WORKBOOK,
        LYQ_SOURCE_ROOT / "标号与参数.xlsx",
        *(LYQ_SOURCE_ROOT / "轮胎" / file_name for _, file_name, *_ in SERVICE_CASES),
        *batch2_required_sources(),
    ]
    missing = [str(path) for path in required if not path.is_file()]
    if missing:
        raise FileNotFoundError("missing experimental asset sources: " + ", ".join(missing))


def _reset_media() -> None:
    if MEDIA_ROOT.exists():
        shutil.rmtree(MEDIA_ROOT)
    (MEDIA_ROOT / "service_tire").mkdir(parents=True)
    (MEDIA_ROOT / "thermal_aging").mkdir(parents=True)


def _build_service_tire_assets() -> list[dict[str, Any]]:
    assets = []
    for sample_id, file_name, tire_model, severity, phenomenon in SERVICE_CASES:
        source = LYQ_SOURCE_ROOT / "轮胎" / file_name
        asset_id = f"LYQ-SERVICE-{sample_id.replace('-', '')}-OPT-01"
        target = MEDIA_ROOT / "service_tire" / f"{asset_id}.jpg"
        _make_thumbnail(source, target)
        assets.append(
            {
                "asset_id": asset_id,
                "experiment_id": "TIRE-LYQ-SERVICE-001",
                "sample_id": f"LYQ-SERVICE-{sample_id}",
                "domain": "tire",
                "topic": "service_tire_wear",
                "title": f"实胎样品 {sample_id} 光学形貌",
                "modality": "optical_image",
                "media_type": "image/jpeg",
                "media_path": target.relative_to(ASSET_ROOT).as_posix(),
                "source_reference": f"授权团队实验资料/LYQ/轮胎/轮胎/{file_name}",
                "source_sha256": _sha256(source),
                "derived_sha256": _sha256(target),
                "condition": {
                    "tire_model": tire_model,
                    "sample_position": sample_id,
                    "evidence_type": "service_tire_surface",
                },
                "labels": [phenomenon, severity],
                "severity": severity,
                "observation": f"样品记录标注为{phenomenon}，等级为{severity}。图像仅用于表面形貌判读训练。",
                "training_use": "比较同型号轮胎不同部位的磨损等级，记录可见特征并提出复检项目。",
                "review_status": "reviewed",
                "authorization_status": "authorized",
                "linked_source_ids": [SERVICE_SOURCE_ID],
                "linked_evidence_ids": [f"LYQ-SERVICE-CASE-{sample_id}"],
                "limitations": "等级来自样品映射表；不得仅凭单张图像作适航放行或报废判断。",
            }
        )
    return assets


def _build_thermal_assets() -> list[dict[str, Any]]:
    index_rows = _sheet_records(RXQ_IMAGE_WORKBOOK, "图像索引")
    annotation_rows = {
        item["image_id"]: item for item in _sheet_records(RXQ_IMAGE_WORKBOOK, "损伤标注")
    }
    case_rows = {
        item["case_id"]: item for item in _sheet_records(RXQ_IMAGE_WORKBOOK, "案例记录")
    }
    assets: list[dict[str, Any]] = []
    for row in index_rows:
        source = TEAM_ROOT / row["relative_path"]
        asset_id = row["image_id"]
        target = MEDIA_ROOT / "thermal_aging" / f"{asset_id}.jpg"
        _make_thumbnail(source, target)
        annotation = annotation_rows.get(asset_id, {})
        case = case_rows.get(row["case_id"], {})
        days = _aging_days(row["sample_id"])
        assets.append(
            {
                "asset_id": asset_id,
                "experiment_id": "TIRE-EXP-RXQ-AGING-001",
                "sample_id": row["sample_id"],
                "domain": "tire",
                "topic": "thermal_aging_multimodal",
                "title": f"热氧老化 {days} d 后{row['image_type']}磨损形貌",
                "modality": "sem_image" if row["image_type"] == "SEM" else "optical_image",
                "media_type": "image/jpeg",
                "media_path": target.relative_to(ASSET_ROOT).as_posix(),
                "source_reference": f"授权团队实验资料/RXQ/{row['relative_path']}",
                "source_sha256": row.get("sha256") or _sha256(source),
                "derived_sha256": _sha256(target),
                "condition": {
                    "aging_temperature_C": 75,
                    "aging_duration_days": days,
                    "normal_load_N": 50,
                    "speed_rpm": 100,
                    "cycles": 10000,
                    "magnification_x": row.get("magnification_x"),
                },
                "labels": str(row.get("damage_type") or "").split("|"),
                "severity": row.get("severity_level") or "training_unrated",
                "observation": annotation.get("evidence_note") or case.get("observed_phenomenon") or "",
                "training_use": case.get("recommended_training_action") or "结合工况和多表征数据进行形貌判读。",
                "review_status": row.get("annotation_status") or "reviewed",
                "authorization_status": row.get("authorization_status") or "authorized",
                "linked_source_ids": THERMAL_SOURCE_IDS,
                "linked_evidence_ids": [row.get("source_record_id", "")],
                "limitations": row.get("notes") or "代表视野不等同于样品总体统计分布。",
            }
        )
    chart_path = MEDIA_ROOT / "thermal_aging" / "RXQ-THERMAL-TREND-01.png"
    chart_summary = _draw_thermal_trend_chart(chart_path)
    assets.insert(
        0,
        {
            "asset_id": "RXQ-THERMAL-TREND-01",
            "experiment_id": "TIRE-EXP-RXQ-AGING-001",
            "sample_id": "RXQ-AGING-GROUP-SUMMARY",
            "domain": "tire",
            "topic": "thermal_aging_multimodal",
            "title": "热氧老化时间与摩擦磨损指标趋势",
            "modality": "derived_curve",
            "media_type": "image/png",
            "media_path": chart_path.relative_to(ASSET_ROOT).as_posix(),
            "source_reference": "授权团队实验资料/RXQ/tire_rxq_aging_experiment_records.xlsx",
            "source_sha256": _sha256(RXQ_EXPERIMENT_WORKBOOK),
            "derived_sha256": _sha256(chart_path),
            "condition": {"aging_temperature_C": 75, **chart_summary},
            "labels": ["热氧老化", "摩擦系数", "质量损失", "界面温度"],
            "severity": "not_applicable",
            "observation": "按老化时长对两次平行试验求均值，用于识别趋势而非拟合寿命模型。",
            "training_use": "要求学习者比较三个指标的变化并指出仅凭趋势不能得出的结论。",
            "review_status": "reviewed",
            "authorization_status": "authorized",
            "linked_source_ids": THERMAL_SOURCE_IDS,
            "linked_evidence_ids": [],
            "limitations": "样本量有限且包含源方案差异标记，不外推为真实服役寿命边界。",
        }
    )
    return assets


def _sheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name]
        headers = [str(value) if value is not None else "" for value in next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))]
        records = []
        for values in sheet.iter_rows(min_row=5, values_only=True):
            if values[0] is None:
                continue
            records.append({key: value for key, value in zip(headers, values) if key})
        return records
    finally:
        workbook.close()


def _draw_thermal_trend_chart(target: Path) -> dict[str, Any]:
    rows = _sheet_records(RXQ_EXPERIMENT_WORKBOOK, "磨损原始记录")
    groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[int(str(row["aging_condition_id"]).split("-")[-1][:-1]) // 24].append(row)
    days = sorted(groups)
    series = {
        "Friction coefficient": [_mean(groups[day], "friction_coefficient") for day in days],
        "Mass loss (mg)": [_mean(groups[day], "mass_loss_mg") for day in days],
        "Surface temp (C)": [_mean(groups[day], "surface_temp_C") for day in days],
    }
    width, height = 1200, 900
    image = Image.new("RGB", (width, height), "#f7faf9")
    draw = ImageDraw.Draw(image)
    colors = ["#176b63", "#b95840", "#315b79"]
    draw.text((50, 28), "Thermo-oxidative aging evidence summary (75 C)", fill="#15272b")
    panel_height = 245
    for index, (label, values) in enumerate(series.items()):
        left, top, right = 90, 100 + index * panel_height, 1140
        bottom = top + 175
        draw.line((left, top, left, bottom), fill="#87989b", width=2)
        draw.line((left, bottom, right, bottom), fill="#87989b", width=2)
        draw.text((left, top - 28), label, fill="#263a3f")
        low, high = min(values), max(values)
        span = high - low or 1.0
        points = []
        for item_index, (day, value) in enumerate(zip(days, values)):
            x = left + item_index * (right - left) / max(1, len(days) - 1)
            y = bottom - 18 - (value - low) / span * (bottom - top - 36)
            points.append((x, y))
            draw.ellipse((x - 6, y - 6, x + 6, y + 6), fill=colors[index])
            draw.text((x - 12, bottom + 8), str(day), fill="#4a5e62")
        if len(points) > 1:
            draw.line(points, fill=colors[index], width=4)
        draw.text((right - 125, top - 28), f"range {low:.3f}-{high:.3f}", fill=colors[index])
    draw.text((90, 850), "Aging duration (days); each point is the mean of two parallel records.", fill="#4a5e62")
    image.save(target, format="PNG", optimize=True)
    return {"aging_days": days, "parallel_records_per_condition": 2}


def _make_thumbnail(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as image:
        rgb = image.convert("RGB")
        rgb.thumbnail((1280, 900))
        canvas = Image.new("RGB", rgb.size, "white")
        canvas.paste(rgb)
        canvas.save(target, format="JPEG", quality=86, optimize=True)


def _write_service_markdown(assets: list[dict[str, Any]]) -> None:
    lines = [
        "---",
        f"source_id: {SERVICE_SOURCE_ID}",
        "domain: tire",
        "source_authority: team_authorized_experiment",
        "authorization_status: authorized",
        "review_status: reviewed",
        "---",
        "",
        "# 航空轮胎实胎偏磨与严重度判读案例",
        "",
        "本专题来自获授权团队实验资料。样品映射表记录了轮胎型号、部位编号、偏磨现象和部分严重度等级；图像仅作为教学判读证据，不构成适航放行或报废结论。",
        "",
        "## 样品与证据",
        "",
        "| 样品 | 轮胎型号 | 标签 | 严重度 | 资产ID |",
        "| --- | --- | --- | --- | --- |",
    ]
    for asset in assets:
        condition = asset["condition"]
        lines.append(
            f"| {asset['sample_id']} | {condition['tire_model']} | {asset['labels'][0]} | "
            f"{asset['severity']} | `{asset['asset_id']}` |"
        )
    lines.extend(
        [
            "",
            "## 判读训练流程",
            "",
            "1. 先核对轮胎型号和样品部位，不跨型号直接比较。",
            "2. 描述磨耗区域分布、方向性纹理、沟槽或裂纹等可见事实，再给出损伤等级。",
            "3. 将图像判断与轮胎压力、起降记录、磨耗深度和无损检查结果交叉验证。",
            "4. 证据不足时输出复检项目，不输出确定的适航放行结论。",
            "",
            "## 常见误区",
            "",
            "- 把局部视野直接当作整条轮胎的统计结论。",
            "- 只凭颜色或亮度判断材料老化程度。",
            "- 忽略同一型号不同部位之间的可比性边界。",
            "- 把教学严重度标签等同于维修手册中的报废限值。",
            "",
        ]
    )
    SERVICE_PARSED_PATH.write_text("\n".join(lines), encoding="utf-8")


def _register_service_source() -> None:
    payload = {"schema_version": "1.0", "sources": []}
    if REGISTRY_PATH.is_file():
        payload = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    entry = {
        "source_id": SERVICE_SOURCE_ID,
        "domain": "tire",
        "title": "航空轮胎实胎偏磨与严重度判读案例",
        "author_or_org": "TEAM-LYQ",
        "document_type": "team_experiment_case_dataset",
        "license_or_authorization": "authorized",
        "local_file": SERVICE_PARSED_PATH.relative_to(CORPUS_ROOT).as_posix(),
        "related_files": [CATALOG_PATH.relative_to(CORPUS_ROOT).as_posix()],
        "source_authority": "team_authorized_experiment",
        "acquisition_status": "team_data_approved",
        "checksum": _sha256(SERVICE_PARSED_PATH),
        "notes": "Ten condition-bound service tire image assets; ungraded samples remain explicitly ungraded.",
    }
    sources = [item for item in payload.get("sources", []) if item.get("source_id") != SERVICE_SOURCE_ID]
    sources.append(entry)
    payload["sources"] = sources
    REGISTRY_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _aging_days(sample_id: str) -> int:
    marker = sample_id.split("-D", 1)[1].split("-", 1)[0]
    return int(marker)


def _mean(rows: list[dict[str, Any]], field: str) -> float:
    return sum(float(item[field]) for item in rows) / len(rows)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


if __name__ == "__main__":
    main()

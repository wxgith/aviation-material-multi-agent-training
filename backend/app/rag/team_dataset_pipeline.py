from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.rag.team_dataset_validator import (
    DEFAULT_DATASET_ROOT,
    EXPERIMENT_TEMPLATE,
    IMAGE_TEMPLATE,
    validate_team_dataset,
)


PROJECT_ROOT = Path(__file__).resolve().parents[3]
CORPUS_ROOT = PROJECT_ROOT / "knowledge_corpus"
PARSED_DIR = CORPUS_ROOT / "parsed_docs"
REGISTRY_PATH = CORPUS_ROOT / "source_registry.json"
PREVIEW_PATH = PROJECT_ROOT / ".tmp" / "team_tire_pipeline_preview.md"

SHEET_PRIMARY_KEYS = {
    "sample_id": "samples",
    "aging_condition_id": "aging_conditions",
    "record_id": "wear_records",
    "image_id": "images",
    "annotation_id": "annotations",
    "case_id": "cases",
}


def load_team_dataset(
    root: Path = DEFAULT_DATASET_ROOT,
    *,
    template_mode: bool = False,
    include_demo: bool = False,
) -> tuple[dict, dict[str, list[dict]]]:
    manifest_path = root / (
        "dataset_manifest.template.json" if template_mode else "dataset_manifest.json"
    )
    if not manifest_path.exists():
        raise FileNotFoundError(f"manifest not found: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    if template_mode:
        workbook_paths = [root / "data" / EXPERIMENT_TEMPLATE, root / "data" / IMAGE_TEMPLATE]
    else:
        files = manifest.get("files") or {}
        workbook_paths = [
            root / str(files.get("experiment_workbook", "")),
            root / str(files.get("image_annotation_workbook", "")),
        ]

    dataset = {name: [] for name in SHEET_PRIMARY_KEYS.values()}
    for workbook_path in workbook_paths:
        if not workbook_path.is_file():
            raise FileNotFoundError(f"workbook not found: {workbook_path}")
        for key, rows in _read_workbook_tables(workbook_path).items():
            dataset[key].extend(rows)

    dataset["wear_records"] = [_derive_wear_fields(row) for row in dataset["wear_records"]]
    if not include_demo:
        dataset = {
            key: [row for row in rows if not _is_demo_record(row)]
            for key, rows in dataset.items()
        }
    return manifest, dataset


def _read_workbook_tables(path: Path) -> dict[str, list[dict]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    tables: dict[str, list[dict]] = {}
    try:
        for worksheet in workbook.worksheets:
            header_row = next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True), ())
            headers = list(header_row)
            first_header = headers[0] if headers else None
            table_name = SHEET_PRIMARY_KEYS.get(str(first_header))
            if not table_name:
                continue
            rows: list[dict] = []
            for values in worksheet.iter_rows(min_row=5, values_only=True):
                record = {
                    str(header): _normalize_value(value)
                    for header, value in zip(headers, values)
                    if header not in (None, "")
                }
                if not record.get(str(first_header)):
                    continue
                rows.append(record)
            tables.setdefault(table_name, []).extend(rows)
    finally:
        workbook.close()
    return tables


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str) and value.startswith("="):
        return None
    return value


def _derive_wear_fields(record: dict) -> dict:
    result = dict(record)
    load = _number(result.get("normal_load_N"))
    speed = _number(result.get("sliding_speed_m_s"))
    duration = _number(result.get("duration_s"))
    friction_force = _number(result.get("friction_force_mean_N"))
    mass_before = _number(result.get("mass_before_mg"))
    mass_after = _number(result.get("mass_after_mg"))

    distance = speed * duration if speed is not None and duration is not None else None
    coefficient = friction_force / load if friction_force is not None and load else None
    mass_loss = mass_before - mass_after if mass_before is not None and mass_after is not None else None
    wear_rate = mass_loss / (load * distance) if mass_loss is not None and load and distance else None
    result.update(
        sliding_distance_m=distance,
        friction_coefficient=coefficient,
        mass_loss_mg=mass_loss,
        wear_rate_mg_per_Nm=wear_rate,
    )
    return result


def _number(value: Any) -> float | None:
    return float(value) if isinstance(value, (int, float)) else None


def _is_demo_record(record: dict) -> bool:
    primary_value = next(iter(record.values()), "")
    return str(primary_value).upper().startswith("DEMO-")


def render_dataset_markdown(manifest: dict, dataset: dict[str, list[dict]]) -> str:
    dataset_id = str(manifest.get("dataset_id") or "team-tire-dataset")
    title = str(manifest.get("title") or "Team tire experiment dataset")
    lines = [
        "---",
        f"source_id: {source_id_from_manifest(manifest)}",
        f"dataset_id: {dataset_id}",
        "domain: tire",
        "source_authority: team_authorized_experiment",
        f"authorization_status: {manifest.get('authorization_status', 'pending')}",
        f"review_status: {(manifest.get('review') or {}).get('review_status', 'draft')}",
        "---",
        "",
        f"# {title}",
        "",
        "## Dataset scope and provenance",
        "",
        f"- Dataset ID: `{dataset_id}`",
        f"- Owner code: `{manifest.get('owner_code', '')}`",
        f"- Authorization: `{manifest.get('authorization_status', 'pending')}`",
        f"- Review: `{(manifest.get('review') or {}).get('review_status', 'draft')}`",
        "- Domain: aviation tire tread rubber damage analysis",
        "",
    ]
    sections = [
        ("Samples", "samples", ["sample_id", "batch_code", "material_type", "aging_condition_id", "authorization_status"]),
        ("Aging conditions", "aging_conditions", ["aging_condition_id", "set_temperature_C", "duration_h", "atmosphere", "qc_flag"]),
        ("Wear records", "wear_records", ["record_id", "sample_id", "normal_load_N", "sliding_speed_m_s", "friction_coefficient", "mass_loss_mg", "wear_rate_mg_per_Nm", "qc_flag"]),
        ("Images", "images", ["image_id", "case_id", "image_type", "damage_type", "severity_level", "reviewer_code"]),
        ("Damage annotations", "annotations", ["annotation_id", "image_id", "damage_label", "confidence", "review_status", "evidence_note"]),
        ("Case records", "cases", ["case_id", "case_title", "observed_phenomenon", "likely_mechanism", "evidence_image_ids", "damage_level", "maintenance_boundary"]),
    ]
    for heading, key, fields in sections:
        lines.extend([f"## {heading}", ""])
        lines.extend(_markdown_table(dataset.get(key, []), fields))
        lines.append("")
    lines.extend(
        [
            "## Evidence and safety boundary",
            "",
            "Each sample, wear record, image, annotation and case identifier is retained as an evidence locator. Generated training content must cite these identifiers and must not convert laboratory observations directly into airworthiness or maintenance conclusions.",
            "",
        ]
    )
    return "\n".join(lines)


def _markdown_table(rows: list[dict], fields: list[str]) -> list[str]:
    if not rows:
        return ["No approved records in this section."]
    output = [
        "| " + " | ".join(fields) + " |",
        "| " + " | ".join("---" for _ in fields) + " |",
    ]
    for row in rows:
        output.append("| " + " | ".join(_markdown_value(row.get(field)) for field in fields) + " |")
    return output


def _markdown_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        text = f"{value:.8g}"
    else:
        text = str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def source_id_from_manifest(manifest: dict) -> str:
    dataset_id = re.sub(r"[^a-z0-9]+", "_", str(manifest.get("dataset_id", "dataset")).lower())
    return f"team_tire_{dataset_id.strip('_')}"


def dry_run_template(root: Path = DEFAULT_DATASET_ROOT) -> dict:
    validation = validate_team_dataset(root)
    if validation["errors"]:
        raise RuntimeError("template package validation failed")
    manifest, dataset = load_team_dataset(root, template_mode=True, include_demo=True)
    markdown = render_dataset_markdown(manifest, dataset)
    PREVIEW_PATH.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW_PATH.write_text(markdown, encoding="utf-8")
    return {
        "status": "preview_only",
        "preview": str(PREVIEW_PATH),
        "record_counts": {key: len(rows) for key, rows in dataset.items()},
        "registered": False,
        "indexed": False,
        "notice": "DEMO records were used only for pipeline verification.",
    }


def ingest_approved_dataset(
    root: Path = DEFAULT_DATASET_ROOT,
    *,
    index: bool = False,
) -> dict:
    validation = validate_team_dataset(root)
    if validation["mode"] != "dataset":
        raise RuntimeError("formal ingestion requires dataset_manifest.json")
    if validation["errors"] or not validation["approved_for_rag"]:
        raise RuntimeError("dataset has not passed authorization and professional review gates")

    manifest, dataset_with_demo = load_team_dataset(root, include_demo=True)
    demo_records = [
        str(next(iter(row.values())))
        for rows in dataset_with_demo.values()
        for row in rows
        if _is_demo_record(row)
    ]
    if demo_records:
        raise RuntimeError(f"formal ingestion rejects DEMO records: {', '.join(demo_records[:5])}")

    source_id = source_id_from_manifest(manifest)
    markdown_path = PARSED_DIR / f"{source_id}.md"
    PARSED_DIR.mkdir(parents=True, exist_ok=True)
    markdown_path.write_text(render_dataset_markdown(manifest, dataset_with_demo), encoding="utf-8")
    _register_source(manifest, source_id, markdown_path, root)

    pipeline_result = None
    if index:
        from app.rag.pipeline import run_pipeline

        pipeline_result = run_pipeline()
    return {
        "status": "ingested",
        "source_id": source_id,
        "parsed_document": str(markdown_path),
        "record_counts": {key: len(rows) for key, rows in dataset_with_demo.items()},
        "registered": True,
        "indexed": bool(index and pipeline_result and pipeline_result.get("indexed_to_es")),
        "pipeline": pipeline_result,
    }


def _register_source(
    manifest: dict,
    source_id: str,
    markdown_path: Path,
    dataset_root: Path,
) -> None:
    payload = {"schema_version": "1.0", "sources": []}
    if REGISTRY_PATH.exists():
        payload = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    files = manifest.get("files") or {}
    experiment_path = dataset_root / str(files.get("experiment_workbook", ""))
    entry = {
        "source_id": source_id,
        "domain": "tire",
        "title": manifest.get("title", source_id),
        "author_or_org": manifest.get("owner_code", "team"),
        "document_type": "team_experiment_dataset",
        "license_or_authorization": manifest.get("authorization_status"),
        "local_file": str(markdown_path.relative_to(CORPUS_ROOT)).replace("\\", "/"),
        "related_files": list((manifest.get("files") or {}).values()),
        "source_authority": "team_authorized_experiment",
        "acquisition_status": "team_data_approved",
        "checksum": _sha256(experiment_path),
        "notes": "Approved team experiment records converted to evidence-addressable Markdown.",
    }
    sources = [item for item in payload.get("sources", []) if item.get("source_id") != source_id]
    sources.append(entry)
    payload["sources"] = sources
    REGISTRY_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert reviewed team tire records into RAG evidence")
    parser.add_argument("--root", type=Path, default=DEFAULT_DATASET_ROOT)
    action = parser.add_mutually_exclusive_group(required=True)
    action.add_argument("--dry-run-template", action="store_true")
    action.add_argument("--ingest", action="store_true")
    parser.add_argument("--index", action="store_true", help="Run the corpus pipeline after ingestion")
    args = parser.parse_args()
    result = (
        dry_run_template(args.root)
        if args.dry_run_template
        else ingest_approved_dataset(args.root, index=args.index)
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

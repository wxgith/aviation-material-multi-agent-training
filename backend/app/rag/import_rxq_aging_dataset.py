from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.rag.team_dataset_validator import EXPERIMENT_TEMPLATE, IMAGE_TEMPLATE


PROJECT_ROOT = Path(__file__).resolve().parents[3]
CORPUS_ROOT = PROJECT_ROOT / "knowledge_corpus"
TEAM_ROOT = CORPUS_ROOT / "raw_docs" / "team_tire"
PENDING_ROOT = TEAM_ROOT / "pending_review" / "rxq_aging_experiments"
DATA_ROOT = PENDING_ROOT / "data"
PARSED_ROOT = CORPUS_ROOT / "parsed_docs"
REGISTRY_PATH = CORPUS_ROOT / "source_registry.json"

DATASET_ID = "TIRE-EXP-RXQ-AGING-001"
SOURCE_ID = "team_tire_tire_exp_rxq_aging_001"
SUPPLEMENT_SOURCE_ID = "team_tire_rxq_aging_multimodal_evidence"
REVIEW_DATE = date(2026, 8, 4).isoformat()

EXPERIMENT_WORKBOOK = "data/tire_rxq_aging_experiment_records.xlsx"
IMAGE_WORKBOOK = "data/tire_rxq_damage_image_annotations.xlsx"
UV_TENSILE_RECORDS = "data/tire_rxq_uv_tensile_records.json"
DESCRIPTION_FILE = "experiment_description_rxq.md"

DAY_SAMPLES = {
    0: [15, 16],
    1: [17, 18],
    2: [21, 22],
    3: [26, 28],
    7: [3, 5],
    14: [41, 42],
    21: [53, 57],
    28: [62, 63],
    35: [80, 78],
}
IMAGE_SAMPLE = {0: 15, 1: 17, 3: 26, 7: 3, 14: 42, 21: 53, 28: 62}
OPTICAL_FILES = {
    0: "15-1.1.jpg",
    1: "17-.1.jpg",
    3: "26-1.1.jpg",
    7: "3-1.1.jpg",
    14: "42-1.1.jpg",
    21: "53-1.1.jpg",
    28: "62-1.jpg",
}
SEM_MAGNIFICATION = {0: 5000, 1: 5000, 3: 5000, 7: 2000, 14: 5000, 21: 5000, 28: 5000}
OBSERVATIONS = {
    0: ("surface_roughening|local_void", "表面起伏与局部孔洞状区域"),
    1: ("abrasion_texture|debris", "平行磨痕纹理及离散颗粒"),
    3: ("periodic_groove|abrasive_wear", "方向性较强的周期沟槽"),
    7: ("local_spalling|debris", "局部片状翘起、剥落与磨屑"),
    14: ("groove|surface_void", "连续沟槽和局部孔洞"),
    21: ("groove|surface_roughening", "连续磨痕沟槽与表面粗化"),
    28: ("debris_agglomeration|groove", "片状磨屑聚集、沟槽与局部材料脱落"),
}

THERMO_AGING_TEMPERATURE_C = 75.0
FRICTION_AMBIENT_TEMPERATURE_C = 25.0
FRICTION_RELATIVE_HUMIDITY_PCT = 45.0
UV_CURVE_ROOT = PENDING_ROOT / "raw_curves" / "uv_tensile"

# The tensile report binds machine sequence numbers to specimen codes. The
# authorized UV protocol then binds those specimen codes to exposure time.
UV_TENSILE_MAPPING = [
    {"csv": "Raw- 9_.csv", "sample": "S23", "hours": 0, "force_N": 401.96, "deformation_mm": 115.28, "strength_MPa": 29.13, "modulus_MPa": 6.92, "elongation_pct": 461.12},
    {"csv": "Raw- 1_.csv", "sample": "S4", "hours": 72, "force_N": 408.63, "deformation_mm": 116.85, "strength_MPa": 29.61, "modulus_MPa": 6.98, "elongation_pct": 467.39},
    {"csv": "Raw- 2_.csv", "sample": "S5", "hours": 144, "force_N": 387.38, "deformation_mm": 108.28, "strength_MPa": 28.07, "modulus_MPa": 6.76, "elongation_pct": 433.11},
    {"csv": "Raw- 3_.csv", "sample": "S8", "hours": 216, "force_N": 427.59, "deformation_mm": 113.59, "strength_MPa": 30.98, "modulus_MPa": 7.67, "elongation_pct": 454.35},
    {"csv": "Raw- 4_.csv", "sample": "S10", "hours": 288, "force_N": 425.26, "deformation_mm": 111.33, "strength_MPa": 30.82, "modulus_MPa": 7.63, "elongation_pct": 445.30},
    {"csv": "Raw- 5_.csv", "sample": "S12", "hours": 360, "force_N": 351.21, "deformation_mm": 91.70, "strength_MPa": 25.45, "modulus_MPa": 7.24, "elongation_pct": 366.80},
    {"csv": "Raw- 6_.csv", "sample": "S14", "hours": 432, "force_N": 400.71, "deformation_mm": 111.58, "strength_MPa": 29.04, "modulus_MPa": 7.25, "elongation_pct": 446.31},
    {"csv": "Raw- 7_.csv", "sample": "S15", "hours": 504, "force_N": 386.80, "deformation_mm": 97.90, "strength_MPa": 28.03, "modulus_MPa": 7.03, "elongation_pct": 391.60},
    {"csv": "Raw- 8_.csv", "sample": "S17", "hours": 576, "force_N": 409.16, "deformation_mm": 116.82, "strength_MPa": 29.65, "modulus_MPa": 7.29, "elongation_pct": 467.27},
]


def main() -> None:
    _assert_sources()
    summaries = _read_primary_summaries()
    experiment_path = _build_experiment_workbook(summaries)
    image_path = _build_image_workbook(summaries)
    uv_records_path = _write_uv_tensile_records(summaries)
    _write_description(summaries)
    supplement_path = _write_supplement(summaries)
    _write_manifest()
    _update_authorization_status()
    _register_supplement(supplement_path)
    print(
        json.dumps(
            {
                "status": "prepared",
                "dataset_id": DATASET_ID,
                "experiment_workbook": str(experiment_path),
                "image_workbook": str(image_path),
                "uv_tensile_records": str(uv_records_path),
                "supplement": str(supplement_path),
                "samples": sum(len(value) for value in DAY_SAMPLES.values()),
                "aging_conditions": len(DAY_SAMPLES),
                "wear_records": len(summaries["wear"]),
                "uv_tensile_curves": len(summaries["uv_tensile"]),
                "images": len(IMAGE_SAMPLE) * 2,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def _assert_sources() -> None:
    required = [
        DATA_ROOT / "原始数据.xlsx",
        DATA_ROOT / "粗糙度.xlsx",
        DATA_ROOT / "EDS元素结果整理.xlsx",
        DATA_ROOT / "DMA结果.xlsx",
        DATA_ROOT / "定伸应力.xlsx",
        TEAM_ROOT / "data" / EXPERIMENT_TEMPLATE,
        TEAM_ROOT / "data" / IMAGE_TEMPLATE,
        *(UV_CURVE_ROOT / item["csv"] for item in UV_TENSILE_MAPPING),
    ]
    missing = [str(path) for path in required if not path.is_file()]
    if missing:
        raise FileNotFoundError("missing required source files: " + ", ".join(missing))


def _read_primary_summaries() -> dict[str, Any]:
    workbook = load_workbook(DATA_ROOT / "原始数据.xlsx", data_only=True, read_only=True)
    try:
        wear = _read_wear_rows(workbook["摩擦试件"])
        friction = _read_paired_summary(workbook["摩擦系数表"])
        temperature = _read_paired_summary(workbook["界面温度"])
    finally:
        workbook.close()
    return {
        "wear": wear,
        "friction": friction,
        "temperature": temperature,
        "roughness": _read_simple_table(DATA_ROOT / "粗糙度.xlsx"),
        "eds": _read_simple_table(DATA_ROOT / "EDS元素结果整理.xlsx"),
        "dma": _read_simple_table(DATA_ROOT / "DMA结果.xlsx"),
        "tensile": _read_simple_table(DATA_ROOT / "定伸应力.xlsx"),
        "contact": _read_simple_table(DATA_ROOT / "印泥法测试.xlsx"),
        "uv_tensile": _read_uv_tensile_curves(),
    }


def _read_uv_tensile_curves() -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for mapping in UV_TENSILE_MAPPING:
        path = UV_CURVE_ROOT / mapping["csv"]
        rows: list[list[float]] = []
        with path.open("r", encoding="gb18030", newline="") as stream:
            reader = csv.reader(stream)
            next(reader, None)
            for values in reader:
                try:
                    rows.append([float(values[index]) for index in range(7)])
                except (IndexError, ValueError):
                    continue
        if not rows:
            raise ValueError(f"UV tensile curve contains no numeric rows: {path}")
        curve = {
            **mapping,
            "row_count": len(rows),
            "raw_max_force_N": max(row[1] for row in rows),
            "raw_max_deformation_mm": max(row[2] for row in rows),
            "raw_max_displacement_mm": max(row[3] for row in rows),
            "raw_max_strain_pct": max(row[5] for row in rows),
            "raw_max_stress_MPa": max(row[6] for row in rows),
        }
        if abs(curve["raw_max_force_N"] - mapping["force_N"]) > 0.02:
            raise ValueError(f"UV tensile force mismatch: {mapping['csv']}")
        if abs(curve["raw_max_stress_MPa"] - mapping["strength_MPa"]) > 0.02:
            raise ValueError(f"UV tensile strength mismatch: {mapping['csv']}")
        output.append(curve)
    return output


def _read_wear_rows(sheet: Any) -> dict[int, dict[str, float]]:
    rows: dict[int, dict[str, float]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if values[0] is None:
            continue
        day = int(values[0])
        sample_raw = values[1]
        sample = 78 if day == 35 and sample_raw == "." else int(sample_raw)
        if sample not in DAY_SAMPLES.get(day, []):
            continue
        rows[sample] = {
            "day": day,
            "mass_before_g": float(values[3]),
            "mass_after_g": float(values[5]),
            "mass_loss_g": float(values[6]),
            "hardness_before": float(values[8]),
            "hardness_after": float(values[10]),
            "hardness_change": float(values[11]),
        }
    expected = {sample for samples in DAY_SAMPLES.values() for sample in samples}
    if set(rows) != expected:
        raise ValueError(f"wear sample mismatch: missing={sorted(expected - set(rows))}")
    return rows


def _read_paired_summary(sheet: Any) -> dict[int, dict[int, float]]:
    output: dict[int, dict[int, float]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not isinstance(values[0], (int, float)):
            continue
        day = int(values[0])
        if day not in DAY_SAMPLES:
            continue
        first_sample = int(values[1])
        second_sample = int(values[3])
        output[day] = {first_sample: float(values[2]), second_sample: float(values[4])}
    return output


def _read_simple_table(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        output = []
        for values in rows:
            if values[0] is None:
                continue
            output.append(
                {header: value for header, value in zip(headers, values) if header}
            )
        return output
    finally:
        workbook.close()


def _build_experiment_workbook(summaries: dict[str, Any]) -> Path:
    target = TEAM_ROOT / EXPERIMENT_WORKBOOK
    shutil.copy2(TEAM_ROOT / "data" / EXPERIMENT_TEMPLATE, target)
    workbook = load_workbook(target)
    samples_sheet = workbook["样品信息"]
    conditions_sheet = workbook["老化工况"]
    wear_sheet = workbook["磨损原始记录"]
    _clear_records(samples_sheet)
    _clear_records(conditions_sheet)
    _clear_records(wear_sheet)

    for day, samples in DAY_SAMPLES.items():
        condition_id = _condition_id(day)
        for repetition, sample in enumerate(samples, start=1):
            samples_sheet.append(
                [
                    _sample_id(day, sample),
                    "RXQ-ANON-BATCH",
                    "航空轮胎胎面橡胶",
                    "ANON-COMPOUND",
                    "矩形块",
                    19,
                    5,
                    8,
                    None,
                    condition_id,
                    repetition,
                    "authorized",
                    f"原试件编号 {sample}；身份与配方信息已脱敏",
                ]
            )
        conditions_sheet.append(
            [
                condition_id,
                THERMO_AGING_TEMPERATURE_C,
                None,
                None,
                day * 24,
                "air",
                "automatic_every_10_s",
                "HD-E701",
                None,
                None,
                len(samples),
                "GB/T 3512-2014",
                "pass_with_source_variant",
                "75 °C 通风老化箱；自动换气每 10 s 一次；0 d 为未暴露对照。源磨损工作簿含 2 d 组，但定稿论文方案未列该组，故保留并标注来源差异。",
            ]
        )

    distance_m = 1256.636
    duration_s = 6000.0
    speed_m_s = distance_m / duration_s
    for day, samples in DAY_SAMPLES.items():
        for repetition, sample in enumerate(samples, start=1):
            record = summaries["wear"][sample]
            coefficient = summaries["friction"][day][sample]
            surface_temp = summaries["temperature"][day][sample]
            wear_sheet.append(
                [
                    _record_id(day, sample),
                    _sample_id(day, sample),
                    _condition_id(day),
                    None,
                    "ANON-OPERATOR",
                    50,
                    speed_m_s,
                    duration_s,
                    distance_m,
                    coefficient * 50,
                    coefficient,
                    record["mass_before_g"] * 1000,
                    record["mass_after_g"] * 1000,
                    record["mass_loss_g"] * 1000,
                    record["mass_loss_g"] * 1_000_000 / (50 * distance_m),
                    surface_temp,
                    FRICTION_AMBIENT_TEMPERATURE_C,
                    FRICTION_RELATIVE_HUMIDITY_PCT,
                    "abrasive_wheel_80_mesh",
                    "CFT-I",
                    repetition,
                    "pending_review/rxq_aging_experiments/data/原始数据.xlsx",
                    "pass_with_documented_conditions",
                    (
                        f"100 r/min、10000 r、50 N；25±1 °C、45±3% RH；80 目砂轮，2000 目砂纸预处理；硬度均值 "
                        f"{record['hardness_before']:.3f}→{record['hardness_after']:.3f} HA。"
                    ),
                ]
            )
    workbook.save(target)
    return target


def _build_image_workbook(summaries: dict[str, Any]) -> Path:
    target = TEAM_ROOT / IMAGE_WORKBOOK
    shutil.copy2(TEAM_ROOT / "data" / IMAGE_TEMPLATE, target)
    workbook = load_workbook(target)
    index_sheet = workbook["图像索引"]
    annotation_sheet = workbook["损伤标注"]
    cases_sheet = workbook["案例记录"]
    _clear_records(index_sheet)
    _clear_records(annotation_sheet)
    _clear_records(cases_sheet)

    for day, sample in IMAGE_SAMPLE.items():
        case_id = f"RXQ-CASE-TIRE-D{day:03d}"
        labels, observed = OBSERVATIONS[day]
        image_ids = []
        for image_type in ("SEM", "optical"):
            image_id = f"RXQ-IMG-D{day:03d}-S{sample}-{image_type.upper()}-01"
            image_ids.append(image_id)
            source_name = f"{sample}-_1.jpeg" if image_type == "SEM" else OPTICAL_FILES[day]
            source_dir = "thermal_sem" if image_type == "SEM" else "thermal_optical"
            extension = ".jpeg" if image_type == "SEM" else ".jpg"
            file_name = f"IMG-RXQ-D{day:03d}-S{sample}-{image_type.upper()}-01{extension}"
            destination_dir = TEAM_ROOT / "images" / image_type
            destination_dir.mkdir(parents=True, exist_ok=True)
            destination = destination_dir / file_name
            shutil.copy2(PENDING_ROOT / "images" / source_dir / source_name, destination)
            relative_path = destination.relative_to(TEAM_ROOT).as_posix()
            index_sheet.append(
                [
                    image_id,
                    case_id,
                    _sample_id(day, sample),
                    file_name,
                    relative_path,
                    image_type,
                    None,
                    "ANON-SEM" if image_type == "SEM" else "ANON-OPTICAL",
                    SEM_MAGNIFICATION[day] if image_type == "SEM" else None,
                    30 if image_type == "SEM" and day != 7 else (80 if image_type == "SEM" else 50),
                    "wear_track_representative_field",
                    labels,
                    "training_unrated",
                    "reviewed",
                    "PROJECT-OWNER",
                    _record_id(day, sample),
                    _sha256(destination),
                    "clean",
                    "authorized",
                    "代表视野；只用于形貌训练，不代表整件样品的统计分布",
                ]
            )
            annotation_sheet.append(
                [
                    f"RXQ-ANN-D{day:03d}-{image_type.upper()}-01",
                    image_id,
                    "whole_image",
                    labels,
                    None,
                    None,
                    None,
                    None,
                    None,
                    0.7,
                    "PROJECT-OWNER",
                    "reviewed",
                    f"可见{observed}；机理判断必须结合摩擦、温度和表征数据",
                    "无像素级定量标注",
                ]
            )
        friction = sum(summaries["friction"][day].values()) / 2
        temperature = sum(summaries["temperature"][day].values()) / 2
        cases_sheet.append(
            [
                case_id,
                f"热氧老化 {day} d 后滑动磨损形貌案例",
                "tire",
                "tread_rubber",
                "50 N、100 r/min、10000 r 的实验室滑动磨损",
                f"{observed}；两次试验平均摩擦系数 {friction:.4f}，平均界面温度 {temperature:.2f} °C",
                "75 °C 热氧老化后的材料状态与滑动剪切共同影响磨痕；仍需结合平行样与多表征证据，不作唯一因果判定",
                "制样差异、局部污染、摩擦副状态、视野选择偏差",
                "|".join(image_ids),
                "|".join(_record_id(day, value) for value in DAY_SAMPLES[day]),
                "training_unrated",
                "对照时序数据并结合 SEM、摩擦系数和温度进行多证据判读",
                "仅用于教学和科研训练，不构成真实航空轮胎维修放行结论",
                "approved",
                "PROJECT-OWNER",
                "authorized",
                "项目负责人确认授权；结论保留实验室边界",
            ]
        )
    workbook.save(target)
    return target


def _write_uv_tensile_records(summaries: dict[str, Any]) -> Path:
    target = TEAM_ROOT / UV_TENSILE_RECORDS
    records = []
    for curve in summaries["uv_tensile"]:
        records.append(
            {
                "evidence_id": f"RXQ-UV-TENSILE-H{curve['hours']:04d}-{curve['sample']}",
                "sample_code": curve["sample"],
                "uv_aging_duration_h": curve["hours"],
                "source_csv": f"raw_curves/uv_tensile/{curve['csv']}",
                "raw_row_count": curve["row_count"],
                "report_metrics": {
                    "maximum_force_N": curve["force_N"],
                    "maximum_deformation_mm": curve["deformation_mm"],
                    "tensile_strength_MPa": curve["strength_MPa"],
                    "elastic_modulus_MPa": curve["modulus_MPa"],
                    "elongation_at_break_pct": curve["elongation_pct"],
                },
                "raw_curve_maxima": {
                    "force_N": curve["raw_max_force_N"],
                    "deformation_mm": curve["raw_max_deformation_mm"],
                    "displacement_mm": curve["raw_max_displacement_mm"],
                    "strain_pct": curve["raw_max_strain_pct"],
                    "stress_MPa": curve["raw_max_stress_MPa"],
                },
            }
        )
    payload = {
        "dataset_id": DATASET_ID,
        "record_type": "uv_aging_tensile_test",
        "authorization_status": "authorized",
        "uv_aging": {
            "equipment": "HZ-2008A",
            "lamp": "UVA-340",
            "wavelength_nm": [295, 365],
            "lamp_max_power_W": 40,
            "irradiance_W_per_m2": 0.6,
        },
        "tensile_test": {
            "equipment": "WDT-10",
            "standard": "GB/T 528",
            "temperature_C": 25,
            "crosshead_speed_mm_per_min": 500,
            "gauge_length_mm": 25,
            "test_date": "2026-04-09",
        },
        "mapping_method": [
            "tensile report machine sequence -> specimen code",
            "authorized UV protocol specimen code -> aging duration",
        ],
        "source_notes": [
            "This supplemental tensile set has no 24 h specimen.",
            "Protocol/files use 504 h; thesis table 2.3 records 507 h. Actual-file value 504 h is retained.",
            "Report deformation/elongation follows the effective fracture endpoint; raw CSV may include post-fracture samples.",
        ],
        "records": records,
    }
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return target


def _write_description(summaries: dict[str, Any]) -> None:
    path = TEAM_ROOT / DESCRIPTION_FILE
    path.write_text(
        "\n".join(
            [
                "# 航空轮胎胎面橡胶热氧/紫外老化与滑动磨损实验说明",
                "",
                "## 数据来源与授权",
                "",
                "本数据集来自团队成员的热氧/紫外老化实验记录。项目负责人已确认可用于本地 RAG、比赛展示和评审提交；人员、配方和设备资产信息采用匿名代码。",
                "",
                "## 热氧老化条件",
                "",
                "- 依据 GB/T 3512-2014；HD-E701 通风老化箱；设定温度 75 °C。",
                "- 空气气氛，样品架转速 1-10 r/min，自动换气每 10 s 一次。",
                "- 定稿论文方案为 0、1、3、7、14、21、28、35 d；源磨损工作簿另含 2 d 组，本数据集按原始记录保留。",
                "",
                "## 摩擦磨损条件",
                "",
                "- 约 19 mm × 5 mm × 8 mm 胎面橡胶块；CFT-I；环-块式滑动。",
                "- 50 N、100 r/min、10000 r、100 min；环境为 25±1 °C、45±3% RH。",
                "- 80 目砂轮，2000 目砂纸预处理；论文报告接触压力约 0.88 MPa。",
                "- 印泥法辅助工作簿中 50 N 对应 62.46 mm²、0.8005 MPa，作为独立测量记录保留。",
                "",
                "## 紫外老化与拉伸条件",
                "",
                "- HZ-2008A，UVA-340 灯管，295-365 nm，最大输出功率 40 W，辐照度 0.6 W/m²。",
                "- WDT-10 拉伸机；GB/T 528；25 °C；500 mm/min；标距 25 mm。",
                "- 本次补做拉伸含 0、72、144、216、288、360、432、504、576 h 九组，不含 24 h。",
                "",
                "## 来源差异与使用边界",
                "",
                "紫外协议、文件夹和拉伸映射均记录 504 h，而论文表 2.3 记为 507 h；本数据按实际文件采用 504 h 并保留差异。实验数据不得直接用于推导适航或维修放行边界。",
                "",
                "## 证据范围",
                "",
                f"正式记录包含 {len(summaries['wear'])} 条热氧磨损记录、{len(summaries['uv_tensile'])} 条紫外拉伸曲线映射、{len(IMAGE_SAMPLE) * 2} 张代表图，以及摩擦系数、界面温度、粗糙度、DMA、EDS 和定伸应力摘要。",
            ]
        ),
        encoding="utf-8",
    )


def _write_supplement(summaries: dict[str, Any]) -> Path:
    PARSED_ROOT.mkdir(parents=True, exist_ok=True)
    path = PARSED_ROOT / f"{SUPPLEMENT_SOURCE_ID}.md"
    lines = [
        "---",
        f"source_id: {SUPPLEMENT_SOURCE_ID}",
        f"dataset_id: {DATASET_ID}",
        "domain: tire",
        "source_authority: team_authorized_experiment",
        "authorization_status: authorized",
        "review_status: approved",
        "---",
        "",
        "# 航空轮胎胎面橡胶老化-磨损多维实验依据",
        "",
        "以下数值均来自已授权实验工作簿。它们用于教学中的趋势识别和多证据判读，不构成航空器维修限值。",
        "",
        "## 热氧老化后摩擦、温度与质量损失",
        "",
        "| 老化/d | 样品 | 摩擦系数 | 界面温度/°C | 质量损失/mg | 硬度变化/HA | 证据ID |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for day, samples in DAY_SAMPLES.items():
        for sample in samples:
            wear = summaries["wear"][sample]
            lines.append(
                f"| {day} | {sample} | {summaries['friction'][day][sample]:.6f} | "
                f"{summaries['temperature'][day][sample]:.5f} | {wear['mass_loss_g'] * 1000:.4f} | "
                f"{wear['hardness_change']:.4f} | `{_record_id(day, sample)}` |"
            )
    lines.extend(_simple_table_section("表面粗糙度 Sa", summaries["roughness"]))
    lines.extend(_simple_table_section("EDS 元素原子分数", summaries["eds"]))
    lines.extend(_simple_table_section("DMA 指标", summaries["dma"]))
    lines.extend(_simple_table_section("定伸应力", summaries["tensile"]))
    lines.extend(_simple_table_section("印泥法接触压力辅助记录", summaries["contact"]))
    lines.extend(
        [
            "",
            "## 可复现实验条件",
            "",
            "热氧老化采用 HD-E701 通风老化箱，75 °C，空气气氛，按 GB/T 3512-2014。滑动磨损采用 CFT-I、50 N、100 r/min、10000 r、100 min，环境为 25±1 °C、45±3% RH；80 目砂轮，2000 目砂纸预处理。",
            "",
            "紫外老化采用 HZ-2008A 与 UVA-340 灯管，295-365 nm，0.6 W/m²。拉伸采用 WDT-10，按 GB/T 528，25 °C、500 mm/min、标距 25 mm。",
            "",
            "## 紫外拉伸样品-时长-曲线映射",
            "",
            "| 老化/h | 样品 | CSV | 报告最大力/N | 抗拉强度/MPa | 弹性模量/MPa | 断裂伸长率/% | 原始曲线最大应力/MPa | 证据ID |",
            "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for curve in summaries["uv_tensile"]:
        evidence_id = f"RXQ-UV-TENSILE-H{curve['hours']:04d}-{curve['sample']}"
        lines.append(
            f"| {curve['hours']} | {curve['sample']} | `{curve['csv']}` | {curve['force_N']:.2f} | "
            f"{curve['strength_MPa']:.2f} | {curve['modulus_MPa']:.2f} | {curve['elongation_pct']:.2f} | "
            f"{curve['raw_max_stress_MPa']:.5f} | `{evidence_id}` |"
        )
    lines.extend(
        [
            "",
            "## 图像证据链",
            "",
            "热氧老化 0、1、3、7、14、21、28 d 各保留一张 SEM 和一张光学代表图。图像、样品和磨损记录通过 RXQ-IMG、RXQ-CASE 与 RXQ-WEAR 编号关联。形貌标签用于训练观察，不应被解释为唯一损伤机理。",
            "",
            "## 来源差异与证据边界",
            "",
            "9 份 CSV 已通过拉伸报告中的机号-样品映射和紫外拉伸协议中的样品-时长映射完成双重绑定。该补做数据不含 24 h 组。协议、文件夹和报告映射采用 504 h，而论文表 2.3 记为 507 h，本版采用实际文件的 504 h 并保留冲突说明。报告的最大变形/伸长率采用有效断裂点，CSV 还可能包含断裂后采样，因此两者同时保留且不相互覆盖。",
            "",
            "热氧源磨损工作簿包含 2 d 组，而定稿论文试验方案未列 2 d；该组按原始记录保留为 source-record variant。所有结果仅用于教学与科研训练，不构成维修放行边界。",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def _simple_table_section(title: str, rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return []
    headers = list(rows[0])
    lines = ["", f"## {title}", "", "| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(_format_value(row.get(header)) for header in headers) + " |")
    return lines


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.8g}"
    return str(value).replace("|", "\\|")


def _write_manifest() -> None:
    payload = {
        "dataset_id": DATASET_ID,
        "title": "航空轮胎胎面橡胶热氧/紫外老化与滑动磨损实验数据集",
        "domain": "tire",
        "owner_code": "TEAM-RXQ",
        "authorization_status": "authorized",
        "license_or_authorization_note": "项目负责人于 2026-08-04 确认可用于系统、比赛展示和提交；公开发布仍采用脱敏版本",
        "contains_personal_data": False,
        "contains_restricted_commercial_data": False,
        "experiment_date_start": "",
        "experiment_date_end": "",
        "material_batch_codes": ["RXQ-ANON-BATCH"],
        "aging_conditions": {
            "thermo_oxidative": {
                "temperature_C": THERMO_AGING_TEMPERATURE_C,
                "durations_d": list(DAY_SAMPLES),
                "standard": "GB/T 3512-2014",
                "equipment": "HD-E701",
            },
            "ultraviolet": {
                "lamp": "UVA-340",
                "irradiance_W_per_m2": 0.6,
                "durations_h": [item["hours"] for item in UV_TENSILE_MAPPING],
                "equipment": "HZ-2008A",
                "source_conflict": "504 h in protocol/files versus 507 h in thesis table 2.3",
            },
        },
        "test_methods": [
            "thermo_oxidative_aging",
            "uv_aging",
            "sliding_wear",
            "optical_morphology",
            "sem",
            "eds",
            "dma",
            "tensile_stress",
        ],
        "files": {
            "experiment_workbook": EXPERIMENT_WORKBOOK,
            "image_annotation_workbook": IMAGE_WORKBOOK,
            "uv_tensile_records": UV_TENSILE_RECORDS,
            "experiment_description": DESCRIPTION_FILE,
            "image_directories": ["images/SEM", "images/optical", "images/damage_cases"],
        },
        "review": {
            "data_reviewer_code": "PROJECT-OWNER",
            "domain_reviewer_code": "PROJECT-OWNER",
            "review_status": "approved",
            "review_date": REVIEW_DATE,
            "scope": "approved for training and competition evidence; not approved as maintenance limits",
        },
        "ingestion": {
            "approved_for_rag": True,
            "source_authority": "team_authorized_experiment",
            "target_index": "aviation_material_knowledge",
            "notes": "热氧与紫外关键工况已由授权论文、协议和报告交叉复核；来源冲突显式保留；生成内容必须引用 evidence IDs 并保留实验室边界",
        },
    }
    (TEAM_ROOT / "dataset_manifest.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _update_authorization_status() -> None:
    payload = {
        "dataset_id": "rxq_aging_experiments",
        "domain": "tire",
        "material": "aircraft_tire_tread_rubber",
        "authorization_status": "authorized",
        "authorized_on": REVIEW_DATE,
        "authorized_by": "PROJECT-OWNER",
        "allowed_for_internal_review": True,
        "allowed_for_rag_index": True,
        "allowed_for_competition_submission": True,
        "allowed_for_public_release": True,
        "conditions": [
            "use anonymized contributor and equipment codes",
            "retain evidence IDs and laboratory-only safety boundary",
            "do not reconstruct or publish proprietary compound formulation",
        ],
    }
    (PENDING_ROOT / "authorization_status.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _register_supplement(path: Path) -> None:
    payload = {"schema_version": "1.0", "sources": []}
    if REGISTRY_PATH.exists():
        payload = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    entry = {
        "source_id": SUPPLEMENT_SOURCE_ID,
        "domain": "tire",
        "title": "航空轮胎胎面橡胶热氧/紫外老化-磨损多维实验依据",
        "author_or_org": "TEAM-RXQ",
        "document_type": "team_experiment_supplement",
        "license_or_authorization": "authorized",
        "local_file": path.relative_to(CORPUS_ROOT).as_posix(),
        "related_files": [
            "raw_docs/team_tire/pending_review/rxq_aging_experiments/data",
            "raw_docs/team_tire/pending_review/rxq_aging_experiments/images",
        ],
        "source_authority": "team_authorized_experiment",
        "acquisition_status": "team_data_approved",
        "checksum": _sha256(DATA_ROOT / "原始数据.xlsx"),
        "notes": "Approved multi-modal experiment summaries; UV curve mapping and aging conditions cross-checked against authorized protocol, report and thesis. Source conflicts remain explicit.",
    }
    sources = [item for item in payload.get("sources", []) if item.get("source_id") != SUPPLEMENT_SOURCE_ID]
    sources.append(entry)
    payload["sources"] = sources
    REGISTRY_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _clear_records(sheet: Any) -> None:
    if sheet.max_row >= 5:
        sheet.delete_rows(5, sheet.max_row - 4)


def _condition_id(day: int) -> str:
    return f"AGE-RXQ-{day * 24:04d}H"


def _sample_id(day: int, sample: int) -> str:
    return f"TIR-RXQ-D{day:03d}-S{sample}"


def _record_id(day: int, sample: int) -> str:
    return f"RXQ-WEAR-D{day:03d}-S{sample}"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


if __name__ == "__main__":
    main()
